import os
import sys
import glob
import pandas as pd
import openpyxl
from datetime import datetime
from dotenv import load_dotenv
from openai import OpenAI

# Add project root to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config.settings import OUTPUT_DIR
from src.utils import logger
from src.rubric_scorer import RubricScorer
import json
import re

# Load environment variables
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), ".env"))

# Initialize OpenAI client
client = OpenAI(
    base_url="https://ollama.com/v1",
    api_key=os.getenv("OLLAMA_API_KEY", "dummy_key")
)
MODEL_NAME = os.getenv("MODEL_NAME", "gpt-oss:20b-cloud")

def get_dynamic_prompt(categories_list):
    categories_str = "\n".join(categories_list)
    return f"""You are an intelligence analyst. Read the following news article and categorize it into EXACTLY ONE of the following buckets:

{categories_str}

Also determine the sentiment of the article as one of: Positive, Negative, or Neutral.
Return ONLY a valid JSON object in this exact format:
{{"category": "Funding", "sentiment": "Positive"}}"""

class IntelligenceEngine:
    """Identifies key competitive actions, threat level, and strategic implications based on LLM category."""
    def __init__(self):
        logger.debug("IntelligenceEngine initialized in categorizer job")

    def extract_insights(self, article_dict):
        category = article_dict.get("News_Category", "General Industry News")
        sentiment = article_dict.get("Sentiment", "Neutral")
        competitor = article_dict.get("Competitor", "Unknown")
        target_brand = article_dict.get("Target_Brand", "Miraee")
        
        is_own_brand = any(brand.lower() in competitor.lower() for brand in ["mondee", "miraee", "abhee", "abhi"])
        
        if is_own_brand:
            if sentiment == "Negative":
                threat = "High"
            else:
                threat = "Low"
        else:
            # Map new categories to threat assessment logic
            if category in ["Partnership and Acquisitions", "Product Announcement"] and sentiment == "Positive":
                threat = "High"
            elif category in ["Funding", "Strategic Expansion or Changes"] and sentiment == "Positive":
                threat = "Medium"
            elif sentiment == "Negative":
                threat = "Low"
            else:
                threat = "Low"
                
        if is_own_brand:
            action = f"Internal corporate {category.lower()} action by {competitor}"
        else:
            action = f"{competitor} executed {category.lower()} movement impacting {target_brand} market segment"
            
        if threat == "High":
            implication = f"High competitive activity. Monitor {competitor}'s {category.lower()} closely and evaluate defense strategies for {target_brand}."
        elif threat == "Medium":
            implication = f"Moderate competitive movement. Assess {target_brand}'s product positioning relative to {competitor}."
        else:
            if sentiment == "Negative" and not is_own_brand:
                implication = f"Potential opportunity. Competitor {competitor} is undergoing stress ({category.lower()}) which {target_brand} can capitalize on."
            else:
                implication = f"Standard market activity by {competitor}. No direct action required for {target_brand}."
            
        return {
            "threat_level": threat,
            "competitor_action": action,
            "strategic_implication": implication
        }


def categorize_article(title, body, scorer):
    text = f"Title: {title}\n\nBody: {body}"
    categories = scorer.get_dynamic_categories()
    prompt = get_dynamic_prompt(categories)
    
    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {"role": "system", "content": prompt},
                {"role": "user", "content": text[:3000]} # Limit text length to avoid token limits
            ],
            temperature=0.1
        )
        content = response.choices[0].message.content.strip()
        
        # Clean markdown tags if present
        if content.startswith("```json"):
            content = content[7:]
        if content.startswith("```"):
            content = content[3:]
        if content.endswith("```"):
            content = content[:-3]
        content = content.strip()
        
        data = json.loads(content)
        category = data.get("category", "General Industry News")
        sentiment = data.get("sentiment", "Neutral")
        
        # Validate category
        valid_found = False
        for valid in categories:
            if valid.lower() in category.lower():
                category = valid
                valid_found = True
                break
        if not valid_found:
            category = "General Industry News"
            
        return category, sentiment
        
    except Exception as e:
        logger.error(f"Error calling LLM or parsing JSON: {e}")
        return "General Industry News", "Neutral"

def get_insights_prompt(brand, filter_type, tab, articles_text):
    tab_focus = {
        "overview": "general corporate threats, strategic competitor movements, and general market developments",
        "growth-marketing": "alliances, joint ventures, funding rounds, capital raises, and corporate restructuring",
        "product-strategy": "product launches, digital releases, feature updates, and technology advancements"
    }[tab]
    
    brand_label = brand.upper()
    
    return f"""You are the Chief Strategy Officer advising the CEO of {brand_label}.
Review the following list of competitor news events from the past {filter_type} relating to {tab_focus}:

{articles_text}

Based on these specific events, write EXACTLY 3 highly actionable, distinct next-step recommendations for our CEO to counter these movements or exploit market opportunities. 
Each recommendation MUST:
1. Be specific to the news presented (do not give generic business advice).
2. Propose a concrete action {brand_label} should take next.
3. Be concise and professional (maximum 2 sentences per recommendation).

Return your response ONLY as a JSON list of strings, like this:
["Recommendation 1...", "Recommendation 2...", "Recommendation 3..."]"""

def generate_ceo_insights(db_file, brand, scorer):
    logger.info(f"Generating synthesized CEO insights for {brand}...")
    try:
        wb = openpyxl.load_workbook(db_file)
        ws = wb["Raw_News"]
        
        # Read worksheet into list of dicts
        header_list = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                rows.append(dict(zip(header_list, row)))
        
        now = datetime.now()
        
        def get_days_old(pub_date):
            if not pub_date:
                return 9999
            try:
                if isinstance(pub_date, str):
                    from dateutil import parser
                    d = parser.parse(pub_date, fuzzy=True).replace(tzinfo=None)
                else:
                    d = pub_date.replace(tzinfo=None)
                return max(0, (now - d).days)
            except Exception:
                return 9999
        
        # Precompute computed days old
        for r in rows:
            r["computed_days_old"] = get_days_old(r.get("Published_Date"))
            
        timeframes = {
            "TODAY": [r for r in rows if r["computed_days_old"] <= 1],
            "7D": [r for r in rows if r["computed_days_old"] <= 7],
            "30D": [r for r in rows if r["computed_days_old"] <= 30],
            "ALL": rows
        }
        
        fallbacks = {
            "overview": [
                "Market conditions are currently stable.",
                "No high-priority competitor moves detected in this timeframe.",
                "Continue tracking standard sector intelligence."
            ],
            "growth-marketing": [
                "No competitor partnerships or alliances reported.",
                "Venture funding and capital movements are currently quiet.",
                "Monitor standard press channels for upcoming deals."
            ],
            "product-strategy": [
                "No competitor product launches or API releases cataloged.",
                "Digital feature velocity remains normal.",
                "Continue monitoring competitor release notes."
            ]
        }
        
        all_results = []
        
        for filter_type, tf_rows in timeframes.items():
            for tab in ["overview", "growth-marketing", "product-strategy"]:
                # Filter rows by tab categories
                if tab == "overview":
                    tab_rows = tf_rows
                elif tab == "growth-marketing":
                    tab_rows = [r for r in tf_rows if r.get("News_Category") in ["Partnership and Acquisitions", "Funding"]]
                else: # product-strategy
                    tab_rows = [r for r in tf_rows if r.get("News_Category") in ["Product Announcement"]]
                
                # Sort by score descending and limit to top 10
                tab_rows = sorted(tab_rows, key=lambda x: int(x.get("Criticality_Score") or 0), reverse=True)[:10]
                
                # Avoid calling LLM if no news is present
                if not tab_rows:
                    insights = fallbacks[tab]
                    logger.info(f"Skipping LLM for {brand} - {filter_type} - {tab} (0 articles). Using fallback.")
                else:
                    # Format articles text
                    articles_text_list = []
                    for r in tab_rows:
                        comp = r.get("Competitor") or "Unknown"
                        title = r.get("Title") or "No Title"
                        cat = r.get("News_Category") or "General"
                        sent = r.get("Sentiment") or "Neutral"
                        score = r.get("Criticality_Score") or 0
                        articles_text_list.append(f"- [{comp}] {title} (Category: {cat}, Sentiment: {sent}, Criticality: {score})")
                    
                    articles_text = "\n".join(articles_text_list)
                    prompt = get_insights_prompt(brand, filter_type, tab, articles_text)
                    
                    try:
                        response = client.chat.completions.create(
                            model=MODEL_NAME,
                            messages=[
                                {"role": "system", "content": prompt}
                            ],
                            temperature=0.1
                        )
                        content = response.choices[0].message.content.strip()
                        
                        # Clean json wrapping
                        if content.startswith("```json"):
                            content = content[7:]
                        if content.startswith("```"):
                            content = content[3:]
                        if content.endswith("```"):
                            content = content[:-3]
                        content = content.strip()
                        
                        bullets = json.loads(content)
                        if isinstance(bullets, list) and len(bullets) >= 3:
                            insights = bullets[:3]
                        else:
                            raise ValueError("Invalid format from LLM")
                    except Exception as e:
                        logger.error(f"Error generating LLM insights for {brand} {filter_type} {tab}: {e}")
                        bullets = re.findall(r'"([^"]+)"', content) if 'content' in locals() else []
                        if len(bullets) >= 3:
                            insights = bullets[:3]
                        else:
                            insights = fallbacks[tab]
                
                # We save exactly 3 insights (pad with empty strings if necessary)
                ins1 = insights[0] if len(insights) > 0 else ""
                ins2 = insights[1] if len(insights) > 1 else ""
                ins3 = insights[2] if len(insights) > 2 else ""
                all_results.append([filter_type, tab, ins1, ins2, ins3])
        
        # Overwrite the sheet entirely
        if "CEO_Insights" in wb.sheetnames:
            del wb["CEO_Insights"]
        
        ws_ins = wb.create_sheet("CEO_Insights")
        ws_ins.append(["Filter_Type", "Tab_ID", "Insight_1", "Insight_2", "Insight_3"])
        for data in all_results:
            ws_ins.append(data)
            
        wb.save(db_file)
        wb.close()
        logger.info(f"Successfully saved CEO insights to {db_file}")
    except Exception as e:
        logger.error(f"Failed to generate CEO insights for {db_file}: {e}")

def run_categorization():
    logger.info("=========================================")
    logger.info("Starting LLM News Categorization Job")
    logger.info("=========================================")
    
    intel_engine = IntelligenceEngine()
    scorer = RubricScorer()
    db_pattern = os.path.join(OUTPUT_DIR, "raw_news_database_*.xlsx")
    db_files = glob.glob(db_pattern)
    
    for db_file in db_files:
        logger.info(f"Processing database: {db_file}")
        
        try:
            wb = openpyxl.load_workbook(db_file)
            ws = wb["Raw_News"]
            
            # Find column indices
            header_list = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            
            # Self-heal historical data: if "Topic" exists instead of "News_Category", rename it
            if "Topic" in header_list and "News_Category" not in header_list:
                idx = header_list.index("Topic")
                ws.cell(row=1, column=idx + 1).value = "News_Category"
                header_list[idx] = "News_Category"
                wb.save(db_file)
                logger.info(f"Renamed legacy 'Topic' column to 'News_Category' in {db_file}")

            try:
                cat_idx = header_list.index("News_Category")
                title_idx = header_list.index("Title")
                body_idx = header_list.index("News_Body")
                comp_idx = header_list.index("Competitor")
                sent_idx = header_list.index("Sentiment")
                target_idx = header_list.index("Target_Brand")
                pub_idx = header_list.index("Published_Date")
                
                # Check for new columns, create if missing
                if "Criticality_Score" not in header_list:
                    ws.cell(row=1, column=len(header_list) + 1).value = "Criticality_Score"
                    header_list.append("Criticality_Score")
                if "UI_Tab_Mapping" not in header_list:
                    ws.cell(row=1, column=len(header_list) + 1).value = "UI_Tab_Mapping"
                    header_list.append("UI_Tab_Mapping")
                    
                crit_idx = header_list.index("Criticality_Score")
                ui_idx = header_list.index("UI_Tab_Mapping")
                
                threat_idx = header_list.index("Threat_Level")
                action_idx = header_list.index("Competitor_Action")
                impl_idx = header_list.index("Strategic_Implication")
            except ValueError as e:
                logger.error(f"Missing required columns in {db_file}: {e}")
                continue
                
            updates_made = 0
            
            for row in ws.iter_rows(min_row=2):
                current_cat = row[cat_idx].value
                
                # Check if cell is empty or NaN-like
                if not current_cat or pd.isna(current_cat) or str(current_cat).strip() == "":
                    title = row[title_idx].value or ""
                    body = row[body_idx].value or ""
                    
                    if not title and not body:
                        continue
                        
                    logger.info(f"Categorizing article: {title[:50]}...")
                    category, sentiment = categorize_article(title, body, scorer)
                    
                    # Update category and sentiment cells
                    row[cat_idx].value = category
                    row[sent_idx].value = sentiment
                    
                    competitor = row[comp_idx].value or ""
                    target_brand = row[target_idx].value or "Miraee"
                    pub_date = row[pub_idx].value
                    
                    is_own_brand = any(b.lower() in competitor.lower() for b in ["mondee", "miraee", "abhee", "abhi"])
                    
                    # Calculate Rubric Score
                    crit_score = scorer.calculate_score(category, sentiment, title, competitor, pub_date, is_own_brand)
                    row[crit_idx].value = crit_score
                    
                    # Calculate UI Tab Mapping
                    ui_tab = scorer.get_ui_mapping(category)
                    row[ui_idx].value = ui_tab
                    
                    # Run Intelligence Engine
                    art_dict = {
                        "News_Category": category,
                        "Competitor": competitor,
                        "Sentiment": sentiment,
                        "Target_Brand": target_brand
                    }
                    insights = intel_engine.extract_insights(art_dict)
                    
                    # Update intelligence cells
                    row[threat_idx].value = insights["threat_level"]
                    row[action_idx].value = insights["competitor_action"]
                    row[impl_idx].value = insights["strategic_implication"]
                    
                    updates_made += 1
                    
            if updates_made > 0:
                wb.save(db_file)
                logger.info(f"Saved {updates_made} updates to {db_file}")
            else:
                logger.info("No uncategorized rows found.")
                
            wb.close()
            
            # Extract brand name from file name
            brand_id = os.path.basename(db_file).split('_')[-1].replace('.xlsx', '').lower()
            
            # Generate or refresh CEO insights sheet in the database
            generate_ceo_insights(db_file, brand_id, scorer)
            
            # Sync the updated file (containing the CEO insights) to the dashboard directory
            import shutil
            dashboard_dir = os.path.join(os.path.dirname(os.path.dirname(OUTPUT_DIR)), "News_Dashboard", "data")
            if os.path.exists(dashboard_dir):
                filename = os.path.basename(db_file)
                dashboard_file_path = os.path.join(dashboard_dir, filename)
                shutil.copy2(db_file, dashboard_file_path)
                logger.info(f"Synced {filename} with CEO insights to dashboard data folder.")
            
        except Exception as e:
            logger.error(f"Error processing {db_file}: {e}")
            
    logger.info("=========================================")
    logger.info("Categorization Job completed.")
    logger.info("=========================================")

if __name__ == "__main__":
    run_categorization()
