import os
import sys
import glob
import openpyxl
import json
from dotenv import load_dotenv
from openai import OpenAI

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.settings import OUTPUT_DIR
from src.utils import logger
from src.rubric_scorer import RubricScorer

load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), ".env"))

client = OpenAI(
    base_url="https://ollama.com/v1",
    api_key=os.getenv("OLLAMA_API_KEY", "dummy_key")
)
MODEL_NAME = os.getenv("MODEL_NAME", "gpt-oss:20b-cloud")

def get_sentiment(title, body):
    text = f"Title: {title}\n\nBody: {body}"
    prompt = """You are an intelligence analyst. Read the following news article and determine the sentiment of the article as one of: Positive, Negative, or Neutral.
Return ONLY a valid JSON object in this exact format:
{"sentiment": "Positive"}"""
    
    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {"role": "system", "content": prompt},
                {"role": "user", "content": text[:3000]}
            ],
            temperature=0.1
        )
        content = response.choices[0].message.content.strip()
        
        if content.startswith("```json"): content = content[7:]
        if content.startswith("```"): content = content[3:]
        if content.endswith("```"): content = content[:-3]
        content = content.strip()
        
        data = json.loads(content)
        return data.get("sentiment", "Neutral")
    except Exception as e:
        logger.error(f"Error calling LLM for sentiment: {e}")
        return "Neutral"

def run_backfill():
    logger.info("Starting Historical Backfill Job for Sentiment and Rubric Scores")
    scorer = RubricScorer()
    
    db_pattern = os.path.join(OUTPUT_DIR, "raw_news_database_*.xlsx")
    for db_file in glob.glob(db_pattern):
        logger.info(f"Processing: {db_file}")
        try:
            wb = openpyxl.load_workbook(db_file)
            ws = wb["Raw_News"]
            header_list = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            
            # Ensure columns exist
            for col in ["Criticality_Score", "UI_Tab_Mapping"]:
                if col not in header_list:
                    ws.cell(row=1, column=len(header_list) + 1).value = col
                    header_list.append(col)
                    
            cat_idx = header_list.index("News_Category")
            title_idx = header_list.index("Title")
            body_idx = header_list.index("News_Body")
            comp_idx = header_list.index("Competitor")
            sent_idx = header_list.index("Sentiment")
            target_idx = header_list.index("Target_Brand")
            pub_idx = header_list.index("Published_Date")
            crit_idx = header_list.index("Criticality_Score")
            ui_idx = header_list.index("UI_Tab_Mapping")
            
            updates = 0
            for row in ws.iter_rows(min_row=2):
                category = row[cat_idx].value
                # Only process historical items that ALREADY have a category
                if category and str(category).strip():
                    title = row[title_idx].value or ""
                    body = row[body_idx].value or ""
                    
                    if not title and not body:
                        continue
                        
                    logger.info(f"Backfilling sentiment for: {title[:50]}...")
                    sentiment = get_sentiment(title, body)
                    row[sent_idx].value = sentiment
                    
                    competitor = row[comp_idx].value or ""
                    pub_date = row[pub_idx].value
                    is_own_brand = any(b.lower() in competitor.lower() for b in ["mondee", "miraee", "abhee", "abhi"])
                    
                    crit_score = scorer.calculate_score(category, sentiment, title, competitor, pub_date, is_own_brand)
                    row[crit_idx].value = crit_score
                    
                    ui_tab = scorer.get_ui_mapping(category)
                    row[ui_idx].value = ui_tab
                    
                    updates += 1
            
            if updates > 0:
                wb.save(db_file)
                logger.info(f"Saved {updates} updates to {db_file}")
                
                import shutil
                dashboard_dir = os.path.join(os.path.dirname(os.path.dirname(OUTPUT_DIR)), "News_Dashboard", "data")
                if os.path.exists(dashboard_dir):
                    filename = os.path.basename(db_file)
                    shutil.copy2(db_file, os.path.join(dashboard_dir, filename))
                    logger.info("Synced to dashboard.")
            wb.close()
            
        except Exception as e:
            logger.error(f"Error processing {db_file}: {e}")
            
    logger.info("Backfill complete.")

if __name__ == "__main__":
    run_backfill()
